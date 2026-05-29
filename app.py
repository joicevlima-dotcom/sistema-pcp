import streamlit as st
import pandas as pd
import pdfplumber
import os
import re
import gspread

from google.oauth2.service_account import Credentials

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Passold - Controle de Saldos",
    layout="wide"
)

# ============================================================
# LOGIN DO SISTEMA
# ============================================================

USUARIOS = {
    "Cenilda": "1234",
    "Joice": "568279",
    "pcp": "2026"
}

if "logado" not in st.session_state:
    st.session_state.logado = False

if "usuario" not in st.session_state:
    st.session_state.usuario = ""

if not st.session_state.logado:

    st.markdown("""
        <div style='text-align:center; padding-top:80px;'>
            <h1 style='color:#1e3a8a;'>🔐 PASSOLD</h1>
            <h3 style='color:#475569;'>Sistema de Gestão de Saldos</h3>
        </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1, 1])

    with col2:
        usuario = st.text_input("Usuário", placeholder="Digite seu usuário")
        senha = st.text_input("Senha", type="password", placeholder="Digite sua senha")

        if st.button("Entrar", use_container_width=True):
            if usuario in USUARIOS and USUARIOS[usuario] == senha:
                st.session_state.logado = True
                st.session_state.usuario = usuario
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")

    st.stop()

# ============================================================
# CSS
# ============================================================

st.markdown("""
    <style>
        .reportview-container { background: #f8fafc; }
        h1, h2, h3 { color: #0f172a; font-family: 'Segoe UI', Arial, sans-serif; }
        .stButton>button {
            background-color: #1e3a8a;
            color: white;
            border-radius: 4px;
            border: none;
            font-weight: 600;
            padding: 0.5rem 2rem;
        }
        .stButton>button:hover { background-color: #1d4ed8; color: white; }
        .stTabs [data-baseweb="tab"] { font-size: 14px; font-weight: 600; color: #475569; }
        .stTabs [data-baseweb="tab"][aria-selected="true"] {
            color: #1e3a8a;
            border-bottom-color: #1e3a8a;
        }
    </style>
""", unsafe_allow_html=True)

# ============================================================
# TOPO
# ============================================================

col_topo1, col_topo2 = st.columns([3, 1])

with col_topo1:
    st.title("Sistema de Gestão de Saldos e Romaneios")
    st.write("Plataforma de Monitoramento e Liberação de Ordens de Produção")

with col_topo2:
    st.write("")
    st.info(f"👤 {st.session_state.usuario}")
    if st.button("🚪 Sair"):
        st.session_state.logado = False
        st.session_state.usuario = ""
        st.rerun()

st.markdown("---")

# ============================================================
# BANCO
# ============================================================

BANCO_DADOS = "banco_ops.xlsx"

# ============================================================
# GOOGLE SHEETS
# ============================================================

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

credenciais = Credentials.from_service_account_info(
    st.secrets["gcp_service_account"],
    scopes=SCOPES
)

client = gspread.authorize(credenciais)

planilha = client.open_by_key("1g6McQiKNTiOHi49hvoGxO86xoeSrM3ktUPLf69hs9bE")
aba_google = planilha.worksheet("ops")
aba_materiais = planilha.worksheet("obras_materiais")
aba_historico = planilha.worksheet("historico_materiais")

# ============================================================
# FUNÇÕES GOOGLE SHEETS
# ============================================================

@st.cache_data(ttl=30)
def carregar_banco():
    try:
        dados = aba_google.get_all_records()

        if dados:
            return pd.DataFrame(dados)

        return pd.DataFrame(columns=[
            "OP", "Obra", "Projeto", "Tipo_Cod",
            "Descricao", "Medida",
            "Qtd_Total", "Qtd_Enviada", "Saldo"
        ])

    except Exception as e:
        st.error(f"Erro ao carregar Google Sheets: {e}")

        return pd.DataFrame(columns=[
            "OP", "Obra", "Projeto", "Tipo_Cod",
            "Descricao", "Medida",
            "Qtd_Total", "Qtd_Enviada", "Saldo"
        ])


def salvar_banco(df):
    try:
        aba_google.clear()

        aba_google.update(
            [df.columns.values.tolist()] +
            df.values.tolist()
        )

        carregar_banco.clear()

    except Exception as e:
        st.error(f"Erro ao salvar Google Sheets: {e}")


@st.cache_data(ttl=30)
def carregar_materiais():
    try:
        dados = aba_materiais.get_all_records()

        if dados:
            return pd.DataFrame(dados)

        return pd.DataFrame(columns=[
            "Obra",
            "Item",
            "Descricao",
            "Unidade",
            "Qtd_Total",
            "Qtd_Enviada",
            "Saldo"
        ])

    except Exception as e:
        st.error(f"Erro ao carregar materiais: {e}")

        return pd.DataFrame(columns=[
            "Obra",
            "Item",
            "Descricao",
            "Unidade",
            "Qtd_Total",
            "Qtd_Enviada",
            "Saldo"
            "Etapa"
        ])


def salvar_materiais(df):
    try:
        aba_materiais.clear()

        aba_materiais.update(
            [df.columns.values.tolist()] +
            df.values.tolist()
        )

        carregar_materiais.clear()

    except Exception as e:
        st.error(f"Erro ao salvar materiais: {e}")

# ============================================================
# TABS
# ============================================================

aba1, aba2, aba3, aba4, aba5 = st.tabs([
    "Importação de O.P. / Cortes / Itens",
    "Emissão de Romaneio Parcial",
    "Painel Geral de Saldos",
    "Acompanhamento de Componentes",
    "Romaneio de Componentes"
])

# ============================================================
# ABA 1: IMPORTAÇÃO INTELIGENTE MULTI-MODELO
# ============================================================

with aba1:

    st.subheader("Registro e Integração de Documentos Técnicos do Projetista")

    col_input1, col_input2, col_input3, col_input4 = st.columns(4)

    with col_input1:
        numero_op = st.text_input("Número da OP:", placeholder="Ex: 1101")

    with col_input2:
        obra_input = st.text_input("Nome da Obra:", placeholder="Ex: PLANETAPEIA")

    with col_input3:
        projeto_input = st.text_input("Projeto/Lote:", placeholder="Ex: PORTAS")

    with col_input4:
        descricao_manual = st.text_input("Descrição Manual:", placeholder="Opcional")

    arquivo_pdf = st.file_uploader("Selecione o PDF (O.P., Padrão de Cortes ou Lista de Itens):", type=["pdf"])

    # ============================================================
    # INCLUSÃO MANUAL DE ITENS
    # ============================================================

    st.markdown("---")

    if "modo_manual" not in st.session_state:
        st.session_state.modo_manual = False

    if st.button("➕ Incluir Item Manualmente"):
        st.session_state.modo_manual = not st.session_state.modo_manual

    if st.session_state.modo_manual:

        st.subheader("Inclusão Manual de Item")

        colm1, colm2, colm3, colm4 = st.columns(4)

        with colm1:
            manual_codigo = st.text_input("Código/Perfil")

        with colm2:
            manual_descricao = st.text_input("Descrição")

        with colm3:
            manual_medida = st.text_input("Medida")

        with colm4:
            manual_qtd = st.number_input(
                "Quantidade",
                min_value=1,
                value=1,
                step=1
            )

        if st.button("💾 Salvar Item Manual"):

            if not numero_op or not obra_input:
                st.error("Preencha o Número da OP e o Nome da Obra antes de salvar.")

            else:

                novo_item = {
                    "OP": str(numero_op).strip(),
                    "Obra": obra_input.upper().strip(),
                    "Projeto": projeto_input.strip(),
                    "Tipo_Cod": manual_codigo.upper().strip(),
                    "Descricao": manual_descricao.upper().strip(),
                    "Medida": manual_medida.strip(),
                    "Qtd_Total": int(manual_qtd),
                    "Qtd_Enviada": 0,
                    "Saldo": int(manual_qtd)
                }

                df_novo = pd.DataFrame([novo_item])

                try:
                    df_existente = carregar_banco()
                    if not df_existente.empty:
                        df_final = pd.concat([df_existente, df_novo], ignore_index=True)
                    else:
                        df_final = df_novo

                    salvar_banco(df_final)

                    st.success("Item manual incluído com sucesso!")
                    st.dataframe(df_novo, use_container_width=True)

                except Exception as e:
                    st.error(f"Erro ao salvar item manual: {e}")

    # ============================================================
    # PROCESSAMENTO DO PDF
    # ============================================================

    if arquivo_pdf is not None:

        st.info("PDF carregado com sucesso.")

        if st.button("🔍 Ver texto bruto do PDF"):
            with pdfplumber.open(arquivo_pdf) as pdf:
                for i, pagina in enumerate(pdf.pages):
                    texto = pagina.extract_text()
                    with st.expander(f"Página {i+1}"):
                        if texto:
                            for j, linha in enumerate(texto.split("\n")):
                                st.code(f"[{j:03d}] {repr(linha)}")
                        else:
                            st.warning("Nenhum texto encontrado.")

        if st.button("Processar Documento Técnico"):
            if not numero_op or not obra_input:
                st.error("Por favor, preencha o Número da OP e o Nome da Obra.")
            else:
                # Identificação inteligente do modelo do arquivo
                tipo_documento = "TRADICIONAL"

                with pdfplumber.open(arquivo_pdf) as pdf:
                    for pagina in pdf.pages:
                        texto_pag = pagina.extract_text()
                        if texto_pag:
                            texto_pag_lower = texto_pag.lower()
                            if "padrões de cortes" in texto_pag_lower or "tam corte" in texto_pag_lower:
                                tipo_documento = "CORTES"
                                break
                            elif "itens da obra" in texto_pag_lower:
                                tipo_documento = "ITENS_OBRA"
                                break

                # =========================================================
                # MODELO 1: RELATÓRIO DE PADRÕES DE CORTES (SmartCEM)
                # =========================================================
                if tipo_documento == "CORTES":
                    st.info("Detectado: Relatório de Padrão de Cortes. Processando barras e ângulos...")
                    itens_extraidos = []

                    perfil_atual = "N/D"
                    descricao_atual = "N/D"
                    estado = "AGUARDANDO"

                    padrao_linha_corte = re.compile(r'^(\d+)\s+(\d+)\s+(\d+(?:[/]\d+)?)')

                    with pdfplumber.open(arquivo_pdf) as pdf:
                        for pagina in pdf.pages:
                            texto = pagina.extract_text()
                            if not texto:
                                continue

                            for linha in texto.split("\n"):
                                linha = linha.strip()
                                if not linha:
                                    continue

                                if "SmartCEM" in linha or "Atenção:" in linha or "Alumisoft" in linha:
                                    continue
                                if linha.startswith("Obra:") and "Calculada em:" in linha:
                                    continue

                                if "Classe/ID:" in linha:
                                    estado = "LER_PERFIL_PROX"
                                    continue

                                if "Qtde total do item" in linha or "Qtde Barra" in linha or "Barra Útil:" in linha:
                                    estado = "IGNORAR_DIAGRAMA"
                                    continue

                                if estado == "IGNORAR_DIAGRAMA":
                                    if re.match(r'^[A-Z]{2,3}-\d{3}', linha):
                                        perfil_atual = linha.split()[0].upper()
                                        estado = "LER_DESCRICAO"
                                    continue

                                if estado == "LER_PERFIL_PROX":
                                    partes_p = linha.split()
                                    if partes_p:
                                        perfil_atual = partes_p[0].upper()
                                    estado = "LER_DESCRICAO"
                                    continue

                                if estado == "LER_DESCRICAO":
                                    if "Nível Otimização" in linha:
                                        descricao_atual = linha.split("Nível Otimização")[0].strip()
                                        estado = "LER_CORTES"
                                        continue
                                    if any(x in linha for x in ["Trat.:", "Part.:", "Nº Barras"]):
                                        continue
                                    if linha.startswith("Qtde") and "Tam" in linha:
                                        estado = "LER_CORTES"
                                        continue
                                    descricao_atual = linha
                                    estado = "LER_CORTES"
                                    continue

                                if estado == "LER_CORTES":
                                    if any(x in linha for x in ["Trat.:", "Part.:", "Nº Barras", "Nível Otimização"]):
                                        continue
                                    if linha.startswith("Qtde") and "Tam" in linha:
                                        continue
                                    if not re.match(r'^\d', linha):
                                        continue

                                    match_corte = padrao_linha_corte.match(linha)
                                    if match_corte:
                                        qtd = int(match_corte.group(1))
                                        tam = int(match_corte.group(2))
                                        corte_tipo = match_corte.group(3).replace('$', '')

                                        if tam in (6000, 6200, 6400):
                                            continue

                                        desc_final = f"{perfil_atual} - {descricao_atual}"
                                        if descricao_manual:
                                            desc_final += f" ({descricao_manual.upper()})"

                                        itens_extraidos.append({
                                            "OP": str(numero_op).strip(),
                                            "Obra": obra_input.upper().strip(),
                                            "Projeto": projeto_input.strip(),
                                            "Tipo_Cod": perfil_atual,
                                            "Descricao": desc_final,
                                            "Medida": f"{tam} mm ({corte_tipo}°)",
                                            "Qtd_Total": qtd,
                                            "Qtd_Enviada": 0,
                                            "Saldo": qtd
                                        })

                # =========================================================
                # MODELO 2: LISTA DE ITENS DA OBRA
                # =========================================================
                elif tipo_documento == "ITENS_OBRA":
                    st.info("Detectado: Lista de Itens da Obra. Processando esquadrias e tipologias...")
                    itens_extraidos = []

                    padrao_item_obra = re.compile(r'^([A-Za-z0-9\-\.]+)\s+(\d+)\s+(\d+)\s+(\d+)')

                    categoria_atual = "Esquadria"

                    with pdfplumber.open(arquivo_pdf) as pdf:
                        for pagina in pdf.pages:
                            texto = pagina.extract_text()
                            if not texto:
                                continue

                            for linha in texto.split("\n"):
                                linha = linha.strip()
                                if not linha or "TOTAIS:" in linha:
                                    continue

                                if ("PORTA" in linha or "JANELA" in linha or "FIXO" in linha or "-" in linha) and len(linha) > 8 and not padrao_item_obra.match(linha):
                                    if "tipo" not in linha.lower() and "quantidade" not in linha.lower() and "obra" not in linha.lower():
                                        categoria_atual = linha.strip()

                                match_item = padrao_item_obra.match(linha)
                                if match_item:
                                    tipo_cod = match_item.group(1).upper()
                                    qtd = int(match_item.group(2))
                                    largura = match_item.group(3)
                                    altura = match_item.group(4)

                                    desc_final = f"{categoria_atual}"
                                    if descricao_manual:
                                        desc_final += f" ({descricao_manual.upper()})"

                                    itens_extraidos.append({
                                        "OP": str(numero_op).strip(),
                                        "Obra": obra_input.upper().strip(),
                                        "Projeto": projeto_input.strip(),
                                        "Tipo_Cod": tipo_cod,
                                        "Descricao": desc_final,
                                        "Medida": f"{largura} x {altura} mm",
                                        "Qtd_Total": qtd,
                                        "Qtd_Enviada": 0,
                                        "Saldo": qtd
                                    })

                # =========================================================
                # MODELO 3: PDF TRADICIONAL DE O.P.
                # =========================================================
                else:
                    st.info("Detectado: PDF Padrão de O.P. Processando...")
                    itens_extraidos = []

                    padrao_op_tradicional = re.compile(r'^([A-Za-z0-9\-\.\/]+)\s+(.*?)\s+(\d+x\d+|\d+[\,\.]?\d*)\s+(\d+)$')

                    with pdfplumber.open(arquivo_pdf) as pdf:
                        for pagina in pdf.pages:
                            texto_op = pagina.extract_text()
                            if not texto_op:
                                continue
                            for linha in texto_op.split("\n"):
                                linha = linha.strip()
                                match_op = padrao_op_tradicional.match(linha)
                                if match_op:
                                    cod = match_op.group(1)
                                    desc = match_op.group(2)
                                    medida = match_op.group(3)
                                    qtd = int(match_op.group(4))

                                    if descricao_manual:
                                        desc = f"{desc} ({descricao_manual.upper()})"

                                    itens_extraidos.append({
                                        "OP": str(numero_op).strip(),
                                        "Obra": obra_input.upper().strip(),
                                        "Projeto": projeto_input.strip(),
                                        "Tipo_Cod": cod,
                                        "Descricao": desc,
                                        "Medida": medida,
                                        "Qtd_Total": qtd,
                                        "Qtd_Enviada": 0,
                                        "Saldo": qtd
                                    })

                # =========================================================
                # SALVAMENTO CENTRALIZADO NO BANCO DE DADOS
                # =========================================================
                if itens_extraidos:
                    df_novos = pd.DataFrame(itens_extraidos)
                    try:
                        df_banco = carregar_banco()
                        if not df_banco.empty and "OP" in df_banco.columns:
                            df_banco = df_banco[df_banco["OP"].astype(str) != str(numero_op).strip()]
                        df_final = pd.concat([df_banco, df_novos], ignore_index=True)
                    except:
                        df_final = df_novos

                    salvar_banco(df_final)
                    st.success(f"Sucesso! {len(df_novos)} itens integrados à base de saldos para a OP {numero_op}.")
                    st.dataframe(df_novos, use_container_width=True)
                else:
                    st.error("Não foi possível extrair dados estruturados deste modelo de arquivo. Valide o alinhamento de texto.")

# ============================================================
# ABA 2: EMISSÃO DE ROMANEIO
# ============================================================

with aba2:

    st.subheader("Ordem de Separação e Carregamento Parcial")

    df_banco = carregar_banco()

    # Inicializa a variável como None para evitar o erro NameError no gerador openpyxl
    imagem_desenho = None

    if not df_banco.empty and "OP" in df_banco.columns:
        ops_disponiveis = df_banco["OP"].unique()
        op_selecionada = st.selectbox("Selecione a OP para Saída de Materiais:", ops_disponiveis)

        itens_op = df_banco[df_banco["OP"].astype(str) == str(op_selecionada)].copy()

        # Atualizado: Criando 3 colunas para acomodar o campo de imagem sem poluir a tela
        col_cab1, col_cab2, col_cab3 = st.columns([1, 2, 2])
        with col_cab1:
            digitado_por = st.text_input("Digitado por:", value="JOICE")
        with col_cab2:
            endereco_obra = st.text_input("Endereço da Obra:")
        with col_cab3:
            # Novo campo de upload para o seu Diretor colocar a foto da travessa/perfil
            imagem_desenho = st.file_uploader("📷 Desenho do Perfil (Opcional):", type=["png", "jpg", "jpeg"])

        st.write("---")
        lista_liberacao = []

        for index, row in itens_op.iterrows():
            col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
            codigo_item = row.get('Tipo_Cod', 'COD')
            descricao_item = row.get('Descricao', 'Sem Descrição')
            measure_item = row.get('Medida', 'Não informada')
            saldo_item = row.get('Saldo', 0)
            qtd_total_item = row.get('Qtd_Total', 0)

            with col1:
                st.write(f"**{codigo_item}** — {descricao_item} ({measure_item})")
            with col2:
                st.write(f"Total: {qtd_total_item} | Saldo: **{saldo_item}**")
            with col3:
                qtd_saida = st.number_input(
                    f"Qtd saída {index}",
                    min_value=0,
                    max_value=int(saldo_item),
                    value=0,
                    key=f"saida_{index}"
                )
                if qtd_saida > 0:
                    lista_liberacao.append({
                        "index": index,
                        "item": row,
                        "qtd_saida": qtd_saida
                    })
            with col4:
                if st.button("🗑️", key=f"del_{index}", help="Excluir este item"):
                    df_banco_del = carregar_banco()
                    df_banco_del = df_banco_del.drop(index=index).reset_index(drop=True)
                    salvar_banco(df_banco_del)
                    st.rerun()

        if len(lista_liberacao) > 0:
            st.markdown("---")
            st.write("### Itens Selecionados para este Romaneio")

            dados_resumo = [
                {
                    "Código/Perfil": lib['item'].get('Tipo_Cod', 'COD'),
                    "Descrição Técnica": lib['item'].get('Descricao', ''),
                    "Dimensão/Corte": lib['item'].get('Medida', ''),
                    "Quantidade": lib['qtd_saida']
                }
                for lib in lista_liberacao
            ]
            st.table(dados_resumo)

            if st.button("Executar Baixa e Emitir Romaneio"):
                df_banco_atual = carregar_banco()

                for lib in lista_liberacao:
                    idx = lib["index"]
                    df_banco_atual.at[idx, "Qtd_Enviada"] = int(df_banco_atual.at[idx, "Qtd_Enviada"]) + lib["qtd_saida"]
                    df_banco_atual.at[idx, "Saldo"] = int(df_banco_atual.at[idx, "Saldo"]) - lib["qtd_saida"]

                salvar_banco(df_banco_atual)
                
                # O seu código do "# Gerador openpyxl" que configuramos entra a partir daqui...
             # Gerador openpyxl
                wb = Workbook()
                ws = wb.active
                ws.sheet_view.showGridLines = False
                ws.title = "Romaneio"

                # Configuração de paginação automática para impressão/PDF
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                ws.oddFooter.right.text = "Página &P de &N"

                bd_fina = Side(style='thin', color="000000")
                borda_padrao = Border(left=bd_fina, right=bd_fina, top=bd_fina, bottom=bd_fina)
                fill_cabecalho = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                # Larguras das colunas baseadas no seu modelo de excelência
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 22
                ws.column_dimensions['C'].width = 48
                ws.column_dimensions['D'].width = 22
                ws.column_dimensions['E'].width = 25

                # LINHAS 1, 2 E 3 - CABEÇALHO (Título e Imagens)
                ws.merge_cells("A1:B3")
                ws.merge_cells("C1:D3")
                ws.merge_cells("E1:E3")

                ws["C1"] = "COMPROVANTE DE ENTREGA DE MATERIAL"
                ws["C1"].font = Font(name="Arial", size=14, bold=True)
                ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

                if os.path.exists("Imagem1.png"):
                    img1 = OpenpyxlImage("Imagem1.png")
                    img1.width = 110
                    img1.height = 45
                    ws.add_image(img1, "A1")

                if os.path.exists("Imagem2.png"):
                    img2 = OpenpyxlImage("Imagem2.png")
                    img2.width = 110
                    img2.height = 45
                    ws.add_image(img2, "E1")

                primeiro_item = lista_liberacao[0]['item']

                # LINHAS 4, 5, 6, 7 E 8 - INFORMAÇÕES GERAIS (TUDO EM MAIÚSCULO)
                ws["A4"] = "Nº:"
                ws["B4"] = str(op_selecionada).upper()
                ws["D4"] = "DIGITADO POR"
                ws["E4"] = str(digitado_por).upper()
                ws["A5"] = "DATA:"
                ws["B5"] = datetime.now().strftime('%d/%m/%Y')
                ws["A6"] = "OBRA:"
                ws["B6"] = str(primeiro_item.get('Obra', '')).upper()
                ws["A7"] = "Nº PROJETO:"
                ws["B7"] = str(primeiro_item.get('Projeto', '')).upper()
                ws["A8"] = "ENDEREÇO DA OBRA:"
                ws["B8"] = str(endereco_obra).upper()

                for r in range(4, 9):
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                    ws.cell(row=r, column=1).font = Font(name="Arial", size=12, bold=True)
                    for c in range(1, 6):
                        ws.cell(row=r, column=c).border = borda_padrao
                        if c > 1:
                            ws.cell(row=r, column=c).font = Font(name="Arial", size=12)

                # LINHA 10 - TÍTULOS DA TABELA
                titulos = ["QTD", "COD / PERFIL", "DESCRIÇÃO TÉCNICA", "MEDIDA / CORTE", "OBSERVAÇÕES"]

                for col_num, titulo in enumerate(titulos, 1):
                    celula = ws.cell(row=10, column=col_num)
                    celula.value = titulo
                    celula.font = Font(name="Arial", size=12, bold=True)
                    celula.fill = fill_cabecalho
                    celula.alignment = Alignment(horizontal="center", vertical="center")
                    celula.border = borda_padrao

                # LINHA 11 - ITENS E DADOS DO CARREGAMENTO
                linha_excel = 11
                for lib in lista_liberacao:
                    item = lib["item"]
                    ws.cell(linha_excel, 1, lib["qtd_saida"])
                    ws.cell(linha_excel, 2, str(item["Tipo_Cod"]).upper())
                    ws.cell(linha_excel, 3, str(item["Descricao"]).upper())
                    ws.cell(linha_excel, 4, str(item["Medida"]).upper())
                    
                    for col in range(1, 6):
                        ws.cell(linha_excel, col).font = Font(name="Arial", size=12)
                        ws.cell(linha_excel, col).border = borda_padrao
                        if col == 3:
                            ws.cell(linha_excel, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        else:
                            ws.cell(linha_excel, col).alignment = Alignment(horizontal="center", vertical="center")
                    linha_excel += 1

                # ============================================================
                # CONTROLE DINÂMICO DO DESENHO E DESLOCAMENTO DO RODAPÉ
                # ============================================================
                deslocamento = 0 
                
                # Se a variável imagem_desenho existir (vinda da tela), coloca o desenho
                if 'imagem_desenho' in locals() and imagem_desenho is not None:
                    try:
                        img_perfil = OpenpyxlImage(imagem_desenho)
                        img_perfil.width = 250   # Ajuste do tamanho do desenho do Diretor
                        img_perfil.height = 150
                        
                        # Define a linha do desenho logo após os itens (com uma linha de folga)
                        linha_foto = max(13, linha_excel + 1)
                        celula_foto = f"C{linha_foto}"
                        ws.add_image(img_perfil, celula_foto)
                        
                        # Empurra o rodapé 9 linhas para baixo para dar espaço perfeito à foto
                        deslocamento = 9 
                    except Exception as e:
                        st.warning(f"Não foi possível carregar o desenho anexado: {e}")

                # Se a tabela cresceu além da linha 11, ajustamos a base do rodapé por segurança
                base_linha = max(0, linha_excel - 12)
                
                # Definição das posições finais com os empurrões inteligentes
                l15 = 15 + deslocamento + base_linha
                l16 = 16 + deslocamento + base_linha
                l17 = 17 + deslocamento + base_linha
                l20 = 20 + deslocamento + base_linha
                l24 = 24 + deslocamento + base_linha
                l27 = 27 + deslocamento + base_linha
                l31 = 31 + deslocamento + base_linha
                l35 = 35 + deslocamento + base_linha

                # LINHAS DE TERMOS E CONDIÇÕES (TAMANHO 14)
                ws.merge_cells(f"A{l15}:E{l15}")
                ws[f"A{l15}"] = "Favor conferir todos os termos descritos neste romaneio antes de assinar."
                ws[f"A{l15}"].font = Font(name="Arial", size=14, italic=True)

                ws.merge_cells(f"A{l16}:E{l16}")
                ws[f"A{l16}"] = "Verificar se os materiais estão em perfeito estado."
                ws[f"A{l16}"].font = Font(name="Arial", size=14, italic=True)

                ws.merge_cells(f"A{l17}:E{l17}")
                ws[f"A{l17}"] = "Não serão aceitas reclamações após recebimento da mercadoria."
                ws[f"A{l17}"].font = Font(name="Arial", size=14, italic=True)

                ws.merge_cells(f"A{l20}:E{l20}")
                ws[f"A{l20}"] = "Recebi da Fachadas Passold as mercadorias acima relacionadas."
                ws[f"A{l20}"].font = Font(name="Arial", size=14)

                # SEÇÃO DE ASSINATURAS MILIMÉTRICAS (Linha em cima, Texto embaixo)
                
                # CONFERÊNCIA INTERNA
                for col in range(1, 4):
                    ws.cell(row=l24, column=col).border = Border(bottom=Side(style='thin'))
                ws[f"D{l24}"] = "____/____/______"
                ws[f"D{l24}"].font = Font(name="Arial", size=12)
                ws[f"A{l24+1}"] = "Conferência Interna:"
                ws[f"A{l24+1}"].font = Font(name="Arial", size=12, bold=True)

                # NOME MOTORISTA
                for col in range(1, 4):
                    ws.cell(row=l27, column=col).border = Border(bottom=Side(style='thin'))
                ws[f"D{l27}"] = "____/____/______"
                ws[f"D{l27}"].font = Font(name="Arial", size=12)
                ws[f"A{l27+1}"] = "Nome Motorista (Data)"
                ws[f"A{l27+1}"].font = Font(name="Arial", size=12, bold=True)

                # NOME RECEBEDOR OBRA
                for col in range(1, 4):
                    ws.cell(row=l31, column=col).border = Border(bottom=Side(style='thin'))
                ws[f"D{l31}"] = "____/____/______"
                ws[f"D{l31}"].font = Font(name="Arial", size=12)
                ws[f"A{l31+1}"] = "Nome Recebedor Obra (Data)"
                ws[f"A{l31+1}"].font = Font(name="Arial", size=12, bold=True)

                # ENGENHEIRO / RESPONSÁVEL OBRA
                for col in range(1, 4):
                    ws.cell(row=l35, column=col).border = Border(bottom=Side(style='thin'))
                ws[f"D{l35}"] = "____/____/______"
                ws[f"D{l35}"].font = Font(name="Arial", size=12)
                ws[f"A{l35+1}"] = "Engenheiro / Responsável Obra"
                ws[f"A{l35+1}"].font = Font(name="Arial", size=12, bold=True)

                # PROCESSAMENTO DO BUFFER E DOWNLOAD NO STREAMLIT
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.success("Romaneio gerado com sucesso! Clique no botão abaixo para baixar.")
                st.download_button(
                    label="⬇️ Baixar Romaneio do Carregamento",
                    data=buffer,
                    file_name=f"Romaneio_OP_{op_selecionada}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
# ============================================================
# ABA 4: ACOMPANHAMENTO LISTA DE MATERIAIS
# ============================================================

with aba4:

    st.subheader("Acompanhamento Geral de Componentes")

    df_materiais = carregar_materiais()

    if not df_materiais.empty:

        obras = sorted(df_materiais["Obra"].dropna().unique())

        obra_selecionada = st.selectbox(
            "Selecione a obra:",
            obras,
            key="obra_componentes"
        )

        st.markdown("---")

        col1, col2 = st.columns(2)

        with col1:
            arquivo_excel = st.file_uploader(
                "Importar Lista de Componentes",
                type=["xlsx"],
                key="upload_componentes"
            )

        with col2:

            if "modo_manual_componentes" not in st.session_state:
                st.session_state.modo_manual_componentes = False

            if st.button("➕ Inserir Item Manualmente"):

                st.session_state.modo_manual_componentes = (
                    not st.session_state.modo_manual_componentes
                )

        # ====================================================
        # FORMULÁRIO MANUAL
        # ====================================================

        if st.session_state.modo_manual_componentes:

            st.markdown("---")

            c1, c2, c3, c4, c5 = st.columns(5)

            with c1:
                item_manual = st.text_input("Item")

            with c2:
                descricao_manual = st.text_input("Descrição")

            with c3:
                unidade_manual = st.text_input("Unidade")

            with c4:
                etapa_manual = st.text_input("Etapa")

            with c5:
                qtd_manual = st.number_input(
                    "Quantidade",
                    min_value=1,
                    value=1
                )

            if st.button("💾 Salvar Componente"):

                novo_item = {
                    "Obra": obra_selecionada,
                    "Item": item_manual,
                    "Descricao": descricao_manual.upper(),
                    "Unidade": unidade_manual.upper(),
                    "Qtd_Total": qtd_manual,
                    "Qtd_Enviada": 0,
                    "Saldo": qtd_manual,
                    "Etapa": etapa_manual.upper(),
                    "Material": ""
                }

                df_novo = pd.DataFrame([novo_item])

                df_final = pd.concat(
                    [df_materiais, df_novo],
                    ignore_index=True
                )

                salvar_materiais(df_final)

                st.success("Componente salvo com sucesso!")

                st.rerun()

        # ====================================================
        # LISTA DA OBRA
        # ====================================================

        st.markdown("---")

        df_obra = df_materiais[
            df_materiais["Obra"] == obra_selecionada
        ]

        st.dataframe(
            df_obra,
            use_container_width=True
        )

    else:

        st.info("Nenhuma lista de componentes cadastrada.")
with aba5:

    st.subheader("Emissão de Romaneio de Componentes")

    df_materiais = carregar_materiais()

    if not df_materiais.empty:

        obras = sorted(
            df_materiais["Obra"].dropna().unique()
        )

        obra_saida = st.selectbox(
            "Selecione a obra:",
            obras,
            key="obra_saida_componentes"
        )

        df_obra = df_materiais[
            df_materiais["Obra"] == obra_saida
        ].copy()

        st.markdown("---")

        col1, col2 = st.columns(2)

        with col1:
            digitado_por_comp = st.text_input(
                "Digitado por:",
                value=st.session_state.usuario
            )

        with col2:
            endereco_comp = st.text_input(
                "Endereço da Obra:"
            )

        lista_saida_componentes = []

        st.markdown("### Componentes Disponíveis")

        for index, row in df_obra.iterrows():

            col_a, col_b, col_c = st.columns([5, 2, 2])

            with col_a:

                st.write(
                    f"""
                    **ITEM {row['Item']}**
                    - {row['Descricao']}
                    - ETAPA: {row.get('Etapa', '')}
                    - MATERIAL: {row.get('Material', '')}
                    """
                )

            with col_b:

                st.write(
                    f"""
                    TOTAL: {row['Qtd_Total']}

                    ENVIADO: {row['Qtd_Enviada']}

                    SALDO: {row['Saldo']}
                    """
                )

            with col_c:

                qtd_saida = st.number_input(
                    f"Saída item {index}",
                    min_value=0,
                    max_value=int(row["Saldo"]),
                    value=0,
                    key=f"comp_saida_{index}"
                )

                if qtd_saida > 0:

                    lista_saida_componentes.append({
                        "index": index,
                        "item": row,
                        "qtd_saida": qtd_saida
                    })

        # ====================================================
        # RESUMO
        # ====================================================

        if len(lista_saida_componentes) > 0:

            st.markdown("---")

            st.write("### Resumo do Romaneio")

            resumo = []

            for item in lista_saida_componentes:

                resumo.append({

                    "Item": item["item"]["Item"],

                    "Descrição": item["item"]["Descricao"],

                    "Quantidade": item["qtd_saida"]
                })

            st.table(pd.DataFrame(resumo))

            # ====================================================
            # BOTÃO EMITIR
            # ====================================================

            if st.button("🚛 Emitir Romaneio Componentes"):

                df_atual = carregar_materiais()

                for item_saida in lista_saida_componentes:

                    idx = item_saida["index"]

                    qtd = item_saida["qtd_saida"]

                    df_atual.at[idx, "Qtd_Enviada"] = (
                        int(df_atual.at[idx, "Qtd_Enviada"]) + qtd
                    )

                    df_atual.at[idx, "Saldo"] = (
                        int(df_atual.at[idx, "Saldo"]) - qtd
                    )

                salvar_materiais(df_atual)

                st.success(
                    "Romaneio de componentes emitido com sucesso!"
                )

                st.rerun()

    else:

        st.info(
            "Nenhuma lista de componentes cadastrada."
        )